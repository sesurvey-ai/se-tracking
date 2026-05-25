"""emcs-api — central API service for pw (Hpw.py) records.

Endpoints:
    GET  /healthz                  → {ok, rows, version}
    GET  /api/records              → list (with filters: since_id, claim_no, limit, offset)
    POST /api/records              → insert single record
    POST /api/records/bulk         → bulk insert (Hpw.py uses this)
    DELETE /api/records/:id        → admin
    GET  /api/records/export.xlsx  → download all
"""
from __future__ import annotations

import io
import logging
import os
from datetime import datetime
from functools import wraps

import orjson
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, request, send_file
from flask_compress import Compress

import db

load_dotenv()
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO").upper(),
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
log = logging.getLogger("emcs-api")

VERSION = "0.1.0"
API_KEY = os.getenv("EMCS_API_KEY", "").strip()

app = Flask(__name__)
Compress(app)


def _jdumps(obj):
    return orjson.dumps(obj).decode()


def require_api_key(fn):
    @wraps(fn)
    def wrapper(*a, **kw):
        if not API_KEY:
            return fn(*a, **kw)
        provided = request.headers.get("X-API-Key") or request.args.get("api_key")
        if provided != API_KEY:
            return jsonify({"error": "unauthorized"}), 401
        return fn(*a, **kw)
    return wrapper


# ---------------------------------------------------------------- health ----

@app.route("/healthz")
def healthz():
    try:
        rows = db.count_records()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True, "version": VERSION, "rows": rows})


# ---------------------------------------------------------------- records ----

_ALL_COLS = (
    "id", "extracted_at",
    "claim_no", "survey_no", "invoice_no", "invoice_seq",
    "date_approve", "offer_amount", "approve_amount", "deduct_amount", "deduct_reason",
    "claim_type", "surveyer", "acc_province",
    "source_file", "source_sheet",
)
_INSERTABLE_COLS = _ALL_COLS[2:]


def _f(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _row_to_tuple(rec: dict) -> tuple | None:
    claim = _s(rec.get("claim_no"))
    if not claim:
        return None
    return (
        claim,
        _s(rec.get("survey_no")),
        _s(rec.get("invoice_no")),
        _s(rec.get("invoice_seq")),
        _s(rec.get("date_approve")),
        _f(rec.get("offer_amount")),
        _f(rec.get("approve_amount")),
        _f(rec.get("deduct_amount")),
        _s(rec.get("deduct_reason")),
        _s(rec.get("claim_type")),
        _s(rec.get("surveyer")),
        _s(rec.get("acc_province")),
        _s(rec.get("source_file")),
        _s(rec.get("source_sheet")),
    )


_UPSERT_SQL = f"""
INSERT INTO pw_records ({", ".join(_INSERTABLE_COLS)})
VALUES ({", ".join(["?"] * len(_INSERTABLE_COLS))})
ON CONFLICT(claim_no, invoice_no, invoice_seq) DO UPDATE SET
    survey_no = excluded.survey_no,
    date_approve = excluded.date_approve,
    offer_amount = excluded.offer_amount,
    approve_amount = excluded.approve_amount,
    deduct_amount = excluded.deduct_amount,
    deduct_reason = excluded.deduct_reason,
    claim_type = excluded.claim_type,
    surveyer = excluded.surveyer,
    acc_province = excluded.acc_province,
    source_file = excluded.source_file,
    source_sheet = excluded.source_sheet,
    extracted_at = datetime('now','localtime')
"""


@app.route("/api/records", methods=["GET"])
@require_api_key
def list_records():
    args = request.args
    where = []
    params: list = []

    if args.get("since_id"):
        where.append("id > ?")
        params.append(int(args["since_id"]))
    if args.get("claim_no"):
        where.append("claim_no = ?")
        params.append(args["claim_no"])
    if args.get("survey_no"):
        where.append("survey_no = ?")
        params.append(args["survey_no"])
    if args.get("invoice_no"):
        where.append("invoice_no = ?")
        params.append(args["invoice_no"])

    limit = max(1, min(int(args.get("limit") or 1000), 10000))
    offset = max(0, int(args.get("offset") or 0))

    where_sql = (" WHERE " + " AND ".join(where)) if where else ""

    conn = db.open_conn()
    try:
        total = conn.execute(f"SELECT COUNT(*) FROM pw_records{where_sql}", params).fetchone()[0]
        rows = conn.execute(
            f"SELECT * FROM pw_records{where_sql} ORDER BY id ASC LIMIT ? OFFSET ?",
            params + [limit, offset],
        ).fetchall()
    finally:
        conn.close()

    return Response(
        _jdumps({
            "rows": [dict(r) for r in rows],
            "total": total, "limit": limit, "offset": offset,
        }),
        mimetype="application/json",
    )


@app.route("/api/records", methods=["POST"])
@require_api_key
def insert_record():
    body = request.get_json(silent=True) or {}
    row = _row_to_tuple(body)
    if not row:
        return jsonify({"error": "claim_no required"}), 400
    with db.txn() as conn:
        cur = conn.execute(_UPSERT_SQL, row)
        rowid = cur.lastrowid
    return jsonify({"ok": True, "id": rowid})


@app.route("/api/records/bulk", methods=["POST"])
@require_api_key
def insert_bulk():
    body = request.get_json(silent=True) or {}
    records = body.get("records") if isinstance(body, dict) else None
    if not isinstance(records, list):
        return jsonify({"error": "expected {records: [...]}"}), 400

    rows: list[tuple] = []
    skipped = 0
    for rec in records:
        t = _row_to_tuple(rec) if isinstance(rec, dict) else None
        if t is None:
            skipped += 1
            continue
        rows.append(t)

    if not rows:
        return jsonify({"ok": True, "inserted": 0, "skipped": skipped})

    with db.txn() as conn:
        BATCH = 1000
        for i in range(0, len(rows), BATCH):
            conn.executemany(_UPSERT_SQL, rows[i:i + BATCH])

    log.info("bulk insert: rows=%d skipped=%d", len(rows), skipped)
    return jsonify({"ok": True, "inserted": len(rows), "skipped": skipped})


@app.route("/api/records/<int:rid>", methods=["DELETE"])
@require_api_key
def delete_record(rid: int):
    with db.txn() as conn:
        cur = conn.execute("DELETE FROM pw_records WHERE id=?", (rid,))
    return jsonify({"ok": True, "deleted": cur.rowcount})


# ---------------------------------------------------------------- export ----

@app.route("/api/records/export.xlsx")
@require_api_key
def export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    conn = db.open_conn()
    try:
        rows = conn.execute(f"SELECT {', '.join(_ALL_COLS)} FROM pw_records ORDER BY id DESC").fetchall()
    finally:
        conn.close()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("pw_records")
    ws.freeze_panes = "A2"
    from openpyxl.cell import WriteOnlyCell

    header = []
    for col in _ALL_COLS:
        c = WriteOnlyCell(ws, value=col)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF1E40AF")
        header.append(c)
    ws.append(header)

    for r in rows:
        ws.append([r[c] for c in _ALL_COLS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return send_file(
        buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"pw_records-{stamp}.xlsx",
    )


# ---------------------------------------------------------------- boot ----

db.init_schema()


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5500"))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
