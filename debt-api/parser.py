"""Parse uploaded Excel files into debt_records rows.

Same column layout as the legacy `extract_ตัดหนี้.py`:
    Column C (index 2) = CLAIM NO.
    Column D (index 3) = เลขที่ใบแจ้งหนี้
    Column F (index 5) = AMT.
Header row search: scan first 5 rows for "เช็ค DD/M/YYYY".
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import IO

import openpyxl

CLAIM_PATTERN = re.compile(r"^\d{4}/[0-9A-Za-z]+$")
DATE_PATTERN = re.compile(r"(\d{1,2}/\d{1,2}/\d{4})")


def _clean_claim(v):
    if v is None:
        return None
    s = str(v).strip()
    if not CLAIM_PATTERN.match(s):
        return None
    return s.replace("/", "")


def _clean_invoice(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.startswith("'"):
        s = s[1:]
    return s or None


def _clean_amount(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_excel(source: bytes | IO[bytes], filename: str | None = None) -> tuple[list[dict], int]:
    """Parse one xlsx blob → (records, skipped_count).

    `source` can be bytes or a file-like object. `filename` is used as the
    `source_file` column on each record.
    """
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(source), data_only=True, read_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True, read_only=True)

    records: list[dict] = []
    skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]

        cut_date = None
        for r in all_rows[:5]:
            for cell in r:
                if cell is None:
                    continue
                s = str(cell)
                if "เช็ค" in s:
                    m = DATE_PATTERN.search(s)
                    if m:
                        cut_date = m.group(1)
                        break
            if cut_date:
                break

        for row in all_rows:
            if len(row) < 6:
                continue
            claim = _clean_claim(row[2])
            invoice = _clean_invoice(row[3])
            amount = _clean_amount(row[5])
            if not claim or not invoice:
                skipped += 1
                continue
            records.append({
                "claim": claim,
                "invoice": invoice,
                "amount": amount,
                "cut_date": cut_date,
                "source_file": filename,
                "sheet": sheet_name,
            })

    wb.close()
    return records, skipped
