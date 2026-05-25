"""debt-api — central API service for ตัดหนี้ records.

Endpoints:
    GET  /healthz                       → liveness + row count
    GET  /api/records                   → list with filters
    POST /api/records/bulk              → bulk insert {records: [...]}
    POST /api/upload                    → multipart upload xlsx, parse, insert
    GET  /api/upload-log                → history of past uploads
    DELETE /api/records/:id             → admin
    GET  /api/records/export.xlsx       → download all
"""
from __future__ import annotations

import io
import logging
import os
from datetime import datetime
from functools import wraps

import orjson
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, request, send_file
from flask_compress import Compress

import db
import parser as xparser

load_dotenv()
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO").upper(),
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
log = logging.getLogger("debt-api")

VERSION = "0.1.0"
API_KEY = os.getenv("DEBT_API_KEY", "").strip()
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "50"))

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
Compress(app)


def _jdumps(obj):
    return orjson.dumps(obj).decode()


def require_api_key(fn):
    @wraps(fn)
    def wrapper(*a, **kw):
        if not API_KEY:
            return fn(*a, **kw)
        provided = request.headers.get("X-API-Key") or request.args.get("api_key")
        if provided != API_KEY:
            return jsonify({"error": "unauthorized"}), 401
        return fn(*a, **kw)
    return wrapper


# ---------------------------------------------------------------- health ----

@app.route("/healthz")
def healthz():
    try:
        rows = db.count_records()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True, "version": VERSION, "rows": rows})


# ---------------------------------------------------------------- helpers ----

_ALL_COLS = ("id", "extracted_at", "claim", "invoice", "amount", "cut_date", "source_file", "sheet")
_INSERTABLE = ("claim", "invoice", "amount", "cut_date", "source_file", "sheet")

_UPSERT_SQL = f"""
INSERT INTO debt_records ({", ".join(_INSERTABLE)})
VALUES ({", ".join(["?"] * len(_INSERTABLE))})
ON CONFLICT(claim, invoice) DO UPDATE SET
    amount      = excluded.amount,
    cut_date    = excluded.cut_date,
    source_file = excluded.source_file,
    sheet       = excluded.sheet,
    extracted_at = datetime('now','localtime')
"""


def _record_tuple(rec: dict) -> tuple | None:
    claim = (rec.get("claim") or "").strip()
    invoice = (rec.get("invoice") or "").strip()
    if not claim or not invoice:
        return None
    amount = rec.get("amount")
    try:
        amount = float(amount) if amount not in (None, "") else None
    except (TypeError, ValueError):
        amount = None
    return (
        claim, invoice, amount,
        (rec.get("cut_date") or None),
        (rec.get("source_file") or None),
        (rec.get("sheet") or None),
    )


def _insert_records(records: list[dict]) -> tuple[int, int, int]:
    """Insert records, returns (added, updated, skipped)."""
    if not records:
        return (0, 0, 0)

    existing = set()
    conn = db.open_conn()
    try:
        for r in conn.execute("SELECT claim, invoice FROM debt_records"):
            existing.add((r["claim"], r["invoice"]))
    finally:
        conn.close()

    rows: list[tuple] = []
    added = updated = skipped = 0
    for rec in records:
        t = _record_tuple(rec)
        if t is None:
            skipped += 1
            continue
        rows.append(t)
        if (t[0], t[1]) in existing:
            updated += 1
        else:
            added += 1
            existing.add((t[0], t[1]))

    with db.txn() as conn:
        BATCH = 1000
        for i in range(0, len(rows), BATCH):
            conn.executemany(_UPSERT_SQL, rows[i:i + BATCH])

    return added, updated, skipped


# ---------------------------------------------------------------- records ----

@app.route("/api/records", methods=["GET"])
@require_api_key
def list_records():
    args = request.args
    where = []
    params: list = []

    if args.get("since_id"):
        where.append("id > ?")
        params.append(int(args["since_id"]))
    if args.get("claim"):
        where.append("claim = ?")
        params.append(args["claim"])
    if args.get("invoice"):
        where.append("invoice = ?")
        params.append(args["invoice"])

    limit = max(1, min(int(args.get("limit") or 1000), 10000))
    offset = max(0, int(args.get("offset") or 0))
    where_sql = (" WHERE " + " AND ".join(where)) if where else ""

    conn = db.open_conn()
    try:
        total = conn.execute(f"SELECT COUNT(*) FROM debt_records{where_sql}", params).fetchone()[0]
        rows = conn.execute(
            f"SELECT * FROM debt_records{where_sql} ORDER BY id ASC LIMIT ? OFFSET ?",
            params + [limit, offset],
        ).fetchall()
    finally:
        conn.close()

    return Response(
        _jdumps({"rows": [dict(r) for r in rows], "total": total, "limit": limit, "offset": offset}),
        mimetype="application/json",
    )


@app.route("/api/records/bulk", methods=["POST"])
@require_api_key
def insert_bulk():
    body = request.get_json(silent=True) or {}
    records = body.get("records") if isinstance(body, dict) else None
    if not isinstance(records, list):
        return jsonify({"error": "expected {records: [...]}"}), 400

    added, updated, skipped = _insert_records(records)
    return jsonify({"ok": True, "added": added, "updated": updated, "skipped": skipped})


@app.route("/api/records/<int:rid>", methods=["DELETE"])
@require_api_key
def delete_record(rid: int):
    with db.txn() as conn:
        cur = conn.execute("DELETE FROM debt_records WHERE id=?", (rid,))
    return jsonify({"ok": True, "deleted": cur.rowcount})


# ---------------------------------------------------------------- upload ----

@app.route("/api/upload", methods=["POST"])
@require_api_key
def upload_xlsx():
    """Multipart upload — field name `file`. Parses xlsx and inserts records."""
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "no file (expected multipart field 'file')"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "only .xlsx files supported"}), 400

    blob = file.read()
    size = len(blob)
    log.info("upload received: %s (%.1f KB)", file.filename, size / 1024)

    upload_id = None
    error_msg = None
    added = updated = skipped = 0
    try:
        records, parser_skipped = xparser.parse_excel(blob, filename=file.filename)
        skipped += parser_skipped
        a, u, s = _insert_records(records)
        added += a
        updated += u
        skipped += s
    except Exception as e:
        error_msg = str(e)
        log.exception("upload parse/insert failed for %s", file.filename)

    # Log the upload regardless
    with db.txn() as conn:
        cur = conn.execute(
            "INSERT INTO upload_log (filename, size_bytes, rows_added, rows_updated, rows_skipped, error) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (file.filename, size, added, updated, skipped, error_msg),
        )
        upload_id = cur.lastrowid

    if error_msg:
        return jsonify({
            "ok": False, "upload_id": upload_id, "filename": file.filename,
            "size_bytes": size, "error": error_msg,
        }), 500

    return jsonify({
        "ok": True, "upload_id": upload_id, "filename": file.filename,
        "size_bytes": size, "added": added, "updated": updated, "skipped": skipped,
    })


@app.route("/api/upload-log")
@require_api_key
def upload_log():
    limit = max(1, min(int(request.args.get("limit") or 50), 500))
    conn = db.open_conn()
    try:
        rows = conn.execute(
            "SELECT * FROM upload_log ORDER BY id DESC LIMIT ?", (limit,),
        ).fetchall()
    finally:
        conn.close()
    return Response(_jdumps({"rows": [dict(r) for r in rows]}), mimetype="application/json")


# ---------------------------------------------------------------- export ----

@app.route("/api/records/export.xlsx")
@require_api_key
def export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.cell import WriteOnlyCell

    conn = db.open_conn()
    try:
        rows = conn.execute(f"SELECT {', '.join(_ALL_COLS)} FROM debt_records ORDER BY id DESC").fetchall()
    finally:
        conn.close()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("debt_records")
    ws.freeze_panes = "A2"
    hdr = []
    for col in _ALL_COLS:
        c = WriteOnlyCell(ws, value=col)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF1E40AF")
        hdr.append(c)
    ws.append(hdr)
    for r in rows:
        ws.append([r[c] for c in _ALL_COLS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return send_file(
        buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"debt_records-{stamp}.xlsx",
    )


# ---------------------------------------------------------------- boot ----

db.init_schema()


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5600"))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
