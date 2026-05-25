"""One-off importer for historical "จบงาน" data exported from old iSurvey.

Reads SE_Report_enquiry_*.xlsx files in ./isurvey/, filters rows where
สถานะงาน == "จบงาน", dedupes by (claim_canon, invoice_canon) keeping the
latest ts, then inserts into stage_closed with negative source_id to mark
them as historical (real se-billing captures have positive id).

Idempotent: deletes existing historical rows (source_id < 0) before insert.

Run:
    python import_isurvey.py
"""
from __future__ import annotations

import logging
import os
import re
import sys
from datetime import datetime

import openpyxl

import db
from normalize import canonical_claim, canonical_invoice

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger("import_isurvey")

ISURVEY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "isurvey")

# Column indices (0-based) — inspected from the 49-col header row
COL_CLAIM         = 0    # เลขเคลม
COL_INVOICE       = 3    # เลขใบแจ้งหนี้ (SEABI-...)
COL_AMPHUR        = 11   # อำเภอที่ออกตรวจสอบ
COL_PROVINCE      = 12   # จังหวัดที่ออกตรวจสอบ
COL_TYPE_CLAIM    = 14   # สด/แห้ง/...
COL_SURVEYOR      = 15   # พนักงานสำรวจ (SEC###)
COL_SERVICE_TYPE  = 22   # ประเภทบริการ (เคลมสด/...)
COL_FINISH_TS     = 32   # วันที่/เวลาเสร็จงาน (ใช้เป็น ts)
COL_TOTAL_AMOUNT  = 38   # รวม (จำนวนเงิน)
COL_INSPECTOR     = 43   # ผู้ตรวจสอบงาน
COL_REVIEW_TS     = 44   # วันที่/เวลาตรวจสอบ
COL_STATUS        = 45   # สถานะงาน — ต้องเป็น "จบงาน"


def _f(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _i(v):
    f = _f(v)
    return int(f) if f is not None else None


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _ts_iso(v):
    """Pass through ISO-ish strings. Excel cells may give datetime objects too."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat(timespec="seconds")
    return str(v).strip() or None


def scan_files(files: list[str]) -> dict:
    """Return dict keyed by (claim_canon, invoice_canon) with latest row."""
    by_key: dict[tuple[str, str], dict] = {}
    total_rows = 0
    total_จบ = 0
    skipped_no_claim = 0

    for path in files:
        fn = os.path.basename(path)
        log.info("opening %s (%.1f MB)", fn, os.path.getsize(path) / 1024 / 1024)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue   # header
            total_rows += 1

            # Filter status
            status = _s(row[COL_STATUS]) if len(row) > COL_STATUS else None
            if status != "จบงาน":
                continue
            total_จบ += 1

            claim_raw = row[COL_CLAIM] if len(row) > COL_CLAIM else None
            invoice_raw = row[COL_INVOICE] if len(row) > COL_INVOICE else None
            claim_canon = canonical_claim(claim_raw)
            if not claim_canon:
                skipped_no_claim += 1
                continue
            invoice_canon = canonical_invoice(invoice_raw) or ""

            ts = _ts_iso(row[COL_FINISH_TS] if len(row) > COL_FINISH_TS else None) \
                 or _ts_iso(row[COL_REVIEW_TS] if len(row) > COL_REVIEW_TS else None)

            new_rec = {
                "claim_canon": claim_canon,
                "invoice_canon": invoice_canon,
                "claim_display": _s(claim_raw) or claim_canon,
                "invoice_display": _s(invoice_raw),
                "ts": ts,
                "amphur_name": _s(row[COL_AMPHUR]) if len(row) > COL_AMPHUR else None,
                "province_name": _s(row[COL_PROVINCE]) if len(row) > COL_PROVINCE else None,
                "surveyor_name": _s(row[COL_SURVEYOR]) if len(row) > COL_SURVEYOR else None,
                "inspector_name": _s(row[COL_INSPECTOR]) if len(row) > COL_INSPECTOR else None,
                "closed_total": _i(row[COL_TOTAL_AMOUNT]) if len(row) > COL_TOTAL_AMOUNT else None,
                "type_claim": _s(row[COL_SERVICE_TYPE]) if len(row) > COL_SERVICE_TYPE else None,
                "source_file": fn,
            }

            key = (claim_canon, invoice_canon)
            existing = by_key.get(key)
            if existing is None:
                by_key[key] = new_rec
            else:
                # keep latest ts (lexicographic ISO compare works)
                if (new_rec["ts"] or "") > (existing["ts"] or ""):
                    by_key[key] = new_rec
        wb.close()
        log.info("  done %s — running unique=%d", fn, len(by_key))

    log.info("TOTALS — rows=%d, จบงาน=%d, no_claim=%d, unique=%d",
             total_rows, total_จบ, skipped_no_claim, len(by_key))
    return by_key


def upsert_to_stage_closed(records: dict) -> tuple[int, int]:
    """Replace all historical (source_id < 0) rows with the given records."""
    synced_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"

    log.info("preparing %d rows for stage_closed (source_id < 0)", len(records))

    rows = []
    for source_id_neg, (key, r) in enumerate(records.items(), start=1):
        rows.append((
            r["claim_canon"], r["invoice_canon"], "", -source_id_neg,
            r["claim_display"], r["invoice_display"], None,   # survey_display
            r["ts"],
            None,                          # province_id
            r["province_name"], r["amphur_name"],
            r["surveyor_name"], r["inspector_name"],
            None, 0,                       # oss_company, is_se
            None,                          # sur_invest
            r["closed_total"],             # ins_invest = total (best-fit)
            None, None, None, None,        # ins_trans, ins_photo, out_of_area_amt, out_of_hours_amt
            None,                          # deduct_amt
            0, 0,                          # late_submit, incomplete_docs
            synced_at,
        ))

    deleted = 0
    inserted = 0
    with db.txn() as conn:
        cur = conn.execute("DELETE FROM stage_closed WHERE source_id < 0")
        deleted = cur.rowcount
        log.info("deleted %d existing historical rows", deleted)

        sql = """
            INSERT INTO stage_closed
                (claim_canonical, invoice_canonical, survey_canonical, source_id,
                 claim_display, invoice_display, survey_display, ts,
                 province_id, province_name, amphur_name,
                 surveyor_name, inspector_name, oss_company, is_se,
                 sur_invest, ins_invest, ins_trans, ins_photo,
                 out_of_area_amt, out_of_hours_amt, deduct_amt,
                 late_submit, incomplete_docs, synced_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        # batch insert for speed
        BATCH = 1000
        for i in range(0, len(rows), BATCH):
            chunk = rows[i:i + BATCH]
            conn.executemany(sql, chunk)
            inserted += len(chunk)
            if (i + BATCH) % 50000 < BATCH:
                log.info("  inserted %d/%d ...", inserted, len(rows))

    log.info("DONE inserted=%d  (deleted=%d historical first)", inserted, deleted)
    return deleted, inserted


def main():
    files = []
    if os.path.isdir(ISURVEY_DIR):
        for f in sorted(os.listdir(ISURVEY_DIR)):
            if f.lower().endswith(".xlsx") and f.startswith("SE_Report"):
                files.append(os.path.join(ISURVEY_DIR, f))
    if not files:
        log.error("no SE_Report_*.xlsx files in %s", ISURVEY_DIR)
        sys.exit(1)
    log.info("files to ingest: %d", len(files))
    for f in files:
        log.info("  - %s", os.path.basename(f))

    db.init_schema()
    by_key = scan_files(files)
    deleted, inserted = upsert_to_stage_closed(by_key)

    # Trigger jobs_view rebuild so the "จบงาน" column updates
    log.info("rebuilding jobs_view (this may take ~1-2 minutes)...")
    import jobs
    written = jobs.rebuild()
    log.info("jobs_view rebuilt: %d rows", written)


if __name__ == "__main__":
    main()
