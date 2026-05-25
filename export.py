"""Streaming Excel export from jobs_view."""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from normalize import thai_be_label

COLUMNS = [
    ("เลขเคลม", "claim_display", 22, "@"),
    ("เลขเซอร์เวย์", "invoice_display", 22, "@"),
    ("สถานะปัจจุบัน", "current_status", 14, None),
    ("จบงาน", "_closed_label", 14, None),
    ("ผู้สำรวจ", "closed_surveyor", 18, None),
    ("จังหวัด", "closed_province", 14, None),
    ("ค่าบริการรวม", "closed_amount", 14, "#,##0"),
    ("บันทึกงาน", "_keyed_label", 14, None),
    ("ผู้คีย์", "keyed_keyer", 18, None),
    ("อนุมัติ", "_approved_label", 14, None),
    ("ยอดอนุมัติ", "approved_amount", 14, "#,##0.00"),
    ("ยอดถูกหัก", "approved_deduct", 14, "#,##0.00"),
    ("ตัดหนี้", "_debt_label", 14, None),
    ("เลขใบแจ้งหนี้", "debt_invoice", 22, "@"),
    ("ยอดตัดหนี้", "debt_amount", 14, "#,##0.00"),
    ("วันที่ตัดหนี้", "_debt_cut_be", 14, None),
    ("ข้ามขั้น", "skipped_stages", 22, None),
]


def _stage_label(flag, iso_date):
    if not flag:
        return "-"
    if iso_date:
        be = thai_be_label(iso_date)
        return f"✓ {be}" if be else "✓"
    return "✓"


def write_xlsx(rows) -> bytes:
    """Build an .xlsx in memory from an iterable of sqlite3.Row jobs_view records."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Tracking")
    ws.freeze_panes = "A2"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1E40AF")

    header_cells = []
    for idx, (label, _key, width, _fmt) in enumerate(COLUMNS, start=1):
        from openpyxl.cell import WriteOnlyCell
        c = WriteOnlyCell(ws, value=label)
        c.font = header_font
        c.fill = header_fill
        header_cells.append(c)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.append(header_cells)

    for row in rows:
        d = dict(row)
        skipped = d.get("skipped_stages")
        if skipped:
            try:
                import json
                skipped_list = json.loads(skipped)
                d["skipped_stages"] = ", ".join(skipped_list) if skipped_list else ""
            except Exception:
                d["skipped_stages"] = ""

        d["_keyed_label"] = _stage_label(d.get("keyed"), d.get("keyed_at"))
        d["_closed_label"] = _stage_label(d.get("closed"), d.get("closed_at"))
        d["_approved_label"] = _stage_label(d.get("approved"), d.get("approved_at"))
        d["_debt_label"] = _stage_label(d.get("debt"), d.get("debt_cut_date"))
        d["_debt_cut_be"] = thai_be_label(d.get("debt_cut_date"))

        out = []
        from openpyxl.cell import WriteOnlyCell
        for label, key, _width, fmt in COLUMNS:
            cell = WriteOnlyCell(ws, value=d.get(key))
            if fmt:
                cell.number_format = fmt
            out.append(cell)
        ws.append(out)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename_now(prefix: str = "tracking") -> str:
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return f"{prefix}-{stamp}.xlsx"
